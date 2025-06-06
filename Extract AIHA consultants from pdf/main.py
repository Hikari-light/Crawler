import csv
from openpyxl import load_workbook

def get_merged_value(ws, row, col):
    for merged in ws.merged_cells.ranges:
        if (row, col) in merged.cells:
            return ws.cell(merged.min_row, merged.min_col).value
    return ws.cell(row, col).value

def get_merged_span(ws, row, col):
    for merged in ws.merged_cells.ranges:
        if (row, col) in merged.cells:
            return merged.max_col - merged.min_col + 1
    return 1

def extract_titles(ws, row, col, max_col=31, expected_titles=12):
    titles = []
    count = 0
    current_col = col

    while current_col <= max_col and count < expected_titles:
        value = get_merged_value(ws, row, current_col)
        if value not in (None, ''):
            titles.append(value)
            span = get_merged_span(ws, row, current_col)
            count += 1
            current_col += span
        else:
            current_col += 1

    return titles, current_col

def is_row_empty(ws, row, max_col):
    return all(
        get_merged_value(ws, row, col) in (None, '') for col in range(1, max_col + 1)
    )

def extract_blocks_and_export_csv(file_path, output_csv):
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    max_row = 380
    max_col = 31  # AE列
    visited_blocks = set()

    all_rows = []
    final_titles = []
    source_link = "https://aiha-assets.sfo2.digitaloceanspaces.com/AIHA/resources/IH_Fit-testingConsultantsList.pdf"

    row = 1
    while row <= max_row:
        col = 1
        while col <= max_col:
            cell_value = get_merged_value(ws, row, col)

            if cell_value not in (None, '') and (row, col) not in visited_blocks:
                print(f"\n📌 检测到数据块起始位置：({row},{col})")
                title, end_col = extract_titles(ws, row, col)
                print("标题：", title)

                if not final_titles:
                    final_titles = ["source_link"] + title

                data_row = row + 1
                while data_row <= max_row and not is_row_empty(ws, data_row, max_col):
                    current_col = col
                    row_data = []

                    while current_col < end_col:
                        val = get_merged_value(ws, data_row, current_col)
                        span = get_merged_span(ws, data_row, current_col)
                        row_data.append(val)
                        current_col += span

                    all_rows.append([source_link] + row_data)

                    for c in range(col, end_col):
                        visited_blocks.add((data_row, c))
                    data_row += 1

                row = data_row
                continue
            col += 1
        row += 1

    # 写入 CSV 文件
    with open(output_csv, mode='w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f, quoting=csv.QUOTE_ALL)
        writer.writerow(final_titles)
        writer.writerows(all_rows)

    print(f"\n✅ 完成导出，共 {len(all_rows)} 条记录，写入：{output_csv}")

# 调用示例
extract_blocks_and_export_csv("scaled2_加水印.xlsx", "output_data.csv")
