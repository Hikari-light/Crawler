from openpyxl import load_workbook

def get_merged_span(ws, row, col):
    """获取单元格横向合并的跨度（最少为1）"""
    for merged in ws.merged_cells.ranges:
        if (row, col) in merged.cells:
            return merged.max_col - merged.min_col + 1
    return 1

def extract_titles(ws, row, col, max_col=31, expected_titles=12):
    """从指定位置开始，向右提取合并标题单元格"""
    titles = []
    count = 0
    current_col = col

    while current_col <= max_col and count < expected_titles:
        value = get_merged_value(ws, row, current_col)
        if value not in (None, ''):
            titles.append(value)
            span = get_merged_span(ws, row, current_col)
            count += 1
            current_col += span  # 跳过合并列
        else:
            current_col += 1  # 忽略空白单元格

    return titles, current_col  # 返回标题 + 实际数据块的宽度终止列

def get_merged_value(ws, row, col):
    """获取合并单元格的值"""
    for merged in ws.merged_cells.ranges:
        if (row, col) in merged.cells:
            return ws.cell(merged.min_row, merged.min_col).value
    return ws.cell(row, col).value

def is_row_empty(ws, row, max_col):
    """判断一整行是否为空"""
    return all(
        get_merged_value(ws, row, col) in (None, '') for col in range(1, max_col + 1)
    )

def extract_blocks_and_print(file_path):
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    max_row = 380
    max_col = 31  # AE列

    visited_blocks = set()

    row = 1
    while row <= max_row:
        col = 1
        while col <= max_col:
            cell_value = get_merged_value(ws, row, col)

            if cell_value not in (None, '') and (row, col) not in visited_blocks:
                print(f"\n📌 检测到数据块起始位置：({row},{col})")

                # 获取标题与实际数据范围
                title, end_col = extract_titles(ws, row, col)
                print("标题：", title)

                # 向下读取数据块
                data_block = []
                data_row = row + 1
                while data_row <= max_row and not is_row_empty(ws, data_row, max_col):
                    current_col = col
                    row_data = []

                    while current_col < end_col:
                        val = get_merged_value(ws, data_row, current_col)
                        span = get_merged_span(ws, data_row, current_col)
                        row_data.append(val)
                        current_col += span  # 跳过合并单元格宽度

                    data_block.append(row_data)

                    # 标记已访问区域（用 title 宽度）
                    for c in range(col, end_col):
                        visited_blocks.add((data_row, c))

                    data_row += 1

                for row_data in data_block:
                    print(row_data)

                row = data_row
                continue

            col += 1
        row += 1


# 示例使用
extract_blocks_and_print("scaled2_加水印.xlsx")



